#!/usr/bin/env python3
"""Build ReviveAI_Sample_Imports.xlsx — one sheet per demo scenario."""

from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

DIR = Path(__file__).parent
CSV_FILES = sorted(DIR.glob("[0-9][0-9]_*.csv"))
OUT = DIR / "ReviveAI_Sample_Imports.xlsx"

wb = Workbook()
wb.remove(wb.active)

for csv_path in CSV_FILES:
    sheet_name = csv_path.stem[:31].replace("_", " ").title()
    ws = wb.create_sheet(title=sheet_name)
    lines = csv_path.read_text(encoding="utf-8").strip().splitlines()
    for row_idx, line in enumerate(lines, start=1):
        for col_idx, cell in enumerate(line.split(","), start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell)

wb.save(OUT)
print(f"Wrote {OUT} ({len(CSV_FILES)} sheets)")
